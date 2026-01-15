import streamlit as st
import numpy as np
import os
import glob
import re
from bs4 import BeautifulSoup
import pandas as pd
import io
import zipfile
from openpyxl.styles import Font, PatternFill, Alignment


# ============================================================
# 1. Caption pattern: Table X. Predicted/Expected Crash ... (Section 1)
# ============================================================
# TABLE_TITLE_PATTERN = re.compile(
#     r"Table\s+\d{1,2}\.\s*"
#     r"(Predicted|Expected)\s+Crash Frequencies and Rates "
#     r"by Highway Segment/Intersection \(Section 1\)", re.I
# )

TABLE_TITLE_PATTERN = re.compile(
    r"Table\s+\d{1,2}\.\s*"
    r"(?:Predicted|Expected)\s+Crash Frequencies and Rates "
    r"by (?:"
        r"Highway Segment/Intersection \(Section 1\)"
        r"|Roundabout\s*\([^)]*\)"
        r"|Ramp Segment/Intersection\s*\([^)]*\)"
        r"|Ramp Terminal\s*\([^)]*\)"
    r")",
    re.I
)



# Column name variants (for robustness across files)
EXPECTED_VARIANTS = {
    "Expected FI": [
        "Expected FI Crash Frequency (crashes/yr)",
        "Expected FI Crash Frequency",
        "Expected FI Crash Frequency (crashes/year)",
    ],
    "Expected PDO": [
        "Expected PDO Crash Frequency (crashes/yr)",
        "Expected PDO Crash Frequency",
        "Expected PDO Crash Frequency (crashes/year)",
    ],
    "Expected Total": [
        "Expected Total Crash Frequency (crashes/yr)",
        "Expected Total Crash Frequency",
        "Expected Total Crash Frequency (crashes/year)",
    ],
}


PREDICTED_VARIANTS = {
    "Predicted FI": [
        "Predicted FI Crash Frequency (crashes/yr)",
        "Predicted FI Crash Frequency",
        "Predicted FI Crash Frequency (crashes/year)",
    ],
    "Predicted PDO": [
        "Predicted PDO Crash Frequency (crashes/yr)",
        "Predicted PDO Crash Frequency",
        "Predicted PDO Crash Frequency (crashes/year)",
    ],
    "Predicted Total": [
        "Predicted Total Crash Frequency (crashes/yr)",
        "Predicted Total Crash Frequency",
        "Predicted Total Crash Frequency (crashes/year)",
    ],
}


# ============================================================
# 2. All helper functions (unchanged)
# ============================================================
def pick_first_present(variants, columns):
    for v in variants:
        if v in columns:
            return v
    return None


def find_section1_table(soup):
    for table in soup.find_all('table'):
        caption = table.find('caption')
        if not caption:
            continue
        caption_text = caption.get_text(strip=True)
        if TABLE_TITLE_PATTERN.search(caption_text):
            return table
    node = soup.find(string=TABLE_TITLE_PATTERN)
    if node:
        parent_table = node.find_parent('table')
        if parent_table:
            return parent_table
    return None


def extract_section1_df(html_path):
    with open(html_path, 'r', encoding='utf-8', errors='ignore') as f:
        html = f.read()
    soup = BeautifulSoup(html, 'html.parser')
    table = find_section1_table(soup)
    if table is None:
        raise RuntimeError("Could not find table with caption 'Table X. Predicted/Expected Crash Frequencies and Rates by Highway Segment/Intersection (Section 1)'.")
    df = pd.read_html(str(table))[0]
    return df


def consolidate_tbl(df):
    df = df.copy()
    df.columns = [re.sub(r'\s+', ' ', str(c).strip()) for c in df.columns]
    cols = list(df.columns)
   
    seg_col = "Segment Number/Intersection Name/Cross Road"
    if seg_col not in cols:
        raise KeyError(f"Segment column '{seg_col}' not found. Available columns: {', '.join(cols)}")
   
    pred_fi_col = pick_first_present(PREDICTED_VARIANTS["Predicted FI"], cols)
    pred_pdo_col = pick_first_present(PREDICTED_VARIANTS["Predicted PDO"], cols)
    pred_tot_col = pick_first_present(PREDICTED_VARIANTS["Predicted Total"], cols)
    missing_pred = []
    for name, col in [("Predicted FI", pred_fi_col), ("Predicted PDO", pred_pdo_col), ("Predicted Total", pred_tot_col)]:
        if col is None:
            missing_pred.append(name)
    if missing_pred:
        raise KeyError(f"Missing required Predicted columns: {', '.join(missing_pred)}. Available headers: {', '.join(cols)}")
   
    exp_fi_col = pick_first_present(EXPECTED_VARIANTS["Expected FI"], cols)
    exp_pdo_col = pick_first_present(EXPECTED_VARIANTS["Expected PDO"], cols)
    exp_tot_col = pick_first_present(EXPECTED_VARIANTS["Expected Total"], cols)
   
    cols_to_keep = [seg_col, pred_fi_col, pred_pdo_col, pred_tot_col]
    if exp_fi_col: cols_to_keep.append(exp_fi_col)
    if exp_pdo_col: cols_to_keep.append(exp_pdo_col)
    if exp_tot_col: cols_to_keep.append(exp_tot_col)
   
    non_numeric_mask = df[seg_col].apply(lambda x: pd.to_numeric(str(x), errors='coerce')).isna()
    filtered_df = df.loc[non_numeric_mask, cols_to_keep].copy()
   
    rename_map = {
        seg_col: "Segment Number/Intersection Name/Cross Road",
        pred_fi_col: "Predicted FI Crash Frequency (crashes/yr)",
        pred_pdo_col: "Predicted PDO Crash Frequency (crashes/yr)",
        pred_tot_col: "Predicted Total Crash Frequency (crashes/yr)",
    }
    if exp_fi_col: rename_map[exp_fi_col] = "Expected FI Crash Frequency (crashes/yr)"
    if exp_pdo_col: rename_map[exp_pdo_col] = "Expected PDO Crash Frequency (crashes/yr)"
    if exp_tot_col: rename_map[exp_tot_col] = "Expected Total Crash Frequency (crashes/yr)"
    filtered_df.rename(columns=rename_map, inplace=True)
   
    if exp_fi_col:
        filtered_df["Expected FI - Predicted FI (crashes/yr)"] = (
            filtered_df["Expected FI Crash Frequency (crashes/yr)"] -
            filtered_df["Predicted FI Crash Frequency (crashes/yr)"]
        )
    if exp_pdo_col:
        filtered_df["Expected PDO - Predicted PDO (crashes/yr)"] = (
            filtered_df["Expected PDO Crash Frequency (crashes/yr)"] -
            filtered_df["Predicted PDO Crash Frequency (crashes/yr)"]
        )
    if exp_tot_col:
        filtered_df["Expected Total - Predicted Total (crashes/yr)"] = (
            filtered_df["Expected Total Crash Frequency (crashes/yr)"] -
            filtered_df["Predicted Total Crash Frequency (crashes/yr)"]
        )
   
    numeric_cols = filtered_df.select_dtypes(include=[np.number]).columns
    filtered_df[numeric_cols] = filtered_df[numeric_cols].round(4)
    return filtered_df


def rename_special_rows(df):
    df = df.copy()
    first_col = df.columns[0]
    row_mapping = {
        'All Segments': 'Total Segment Crashes',
        'All Intersections': 'Total Intersection Crashes',
        'Total': 'Total',
    }
    mask = df[first_col].isin(row_mapping.keys())
    df.loc[mask, first_col] = df.loc[mask, first_col].map(row_mapping)
    return df


def update_summary_tables(filename, consolidated_df, inter_summary, seg_summary):
    df = consolidated_df.copy()
    first_col = df.columns[0]
    seg_total_label = 'Total Segment Crashes'
    int_total_label = 'Total Intersection Crashes'
    corridor_total_label = 'Total'
   
    if len(df) > 1 or df[first_col].iloc[0] != corridor_total_label:
        mask_inter = df[first_col].isin([seg_total_label, int_total_label, corridor_total_label])
        inter_rows = df.loc[~mask_inter].copy()
        if not inter_rows.empty:
            inter_rows.insert(0, 'Source File', filename)
            inter_summary = pd.concat([inter_summary, inter_rows], ignore_index=True)
         
        mask_seg = df[first_col] == seg_total_label
        seg_rows = df.loc[mask_seg].copy()
        if not seg_rows.empty:
            seg_rows.insert(0, 'Source File', filename)
            seg_summary = pd.concat([seg_summary, seg_rows], ignore_index=True)
    else:
        total_row = df.copy()
        total_row.insert(0, 'Source File', filename)
        seg_summary = pd.concat([seg_summary, total_row], ignore_index=True)
   
    return inter_summary, seg_summary


def append_total_row(df, id_col_name='Source File'):
    if df.empty:
        return df
    df = df.copy()
    total_row = {}
    total_row[id_col_name] = 'Total'
    for col in df.columns:
        if col == id_col_name:
            continue
        if df[col].dtype in [np.int64, np.int32, np.float64, np.float32, 'int64', 'int32', 'float64', 'float32']:
            total_row[col] = df[col].sum()
        else:
            total_row[col] = ''
    total_df = pd.DataFrame([total_row])
    df = pd.concat([df, total_df], ignore_index=True)
    return df


def create_individual_excel(raw_df, consolidated_df, filename):
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        raw_df.to_excel(writer, sheet_name='Original', index=False)
        consolidated_df.to_excel(writer, sheet_name='Consolidated', index=False)
         
        workbook = writer.book
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
         
        for sheet in workbook.sheetnames:
            worksheet = workbook[sheet]
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
           
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
   
    buffer.seek(0)
    return buffer


def create_batch_zip(results_data, inter_summary, seg_summary):
    zip_buffer = io.BytesIO()
   
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
        # 1. Add ALL individual Excel files first
        for filename, data in results_data.items():
            excel_buffer = create_individual_excel(data['raw'], data['consolidated'], filename)
            clean_filename = filename.replace('.html', '').replace('.htm', '')
            zf.writestr(f"Individual Files/{clean_filename}_extracted.xlsx", excel_buffer.getvalue())
       
        # 2. Create ONLY ONE summary file
        summary_buffer = io.BytesIO()
        sheets_written = False
        with pd.ExcelWriter(summary_buffer, engine='openpyxl') as writer:
            if not inter_summary.empty:
                inter_with_total = append_total_row(inter_summary, 'Source File')
                inter_with_total.to_excel(writer, sheet_name='Intersection Crashes', index=False)
                sheets_written = True
            if not seg_summary.empty:
                seg_with_total = append_total_row(seg_summary, 'Source File')
                seg_with_total.to_excel(writer, sheet_name='Segment Crashes', index=False)
                sheets_written = True
             
            if not sheets_written:
                pd.DataFrame({'Message': ['No intersection/segment data found.']}).to_excel(writer, sheet_name='Summary', index=False)


        summary_buffer.seek(0)
        zf.writestr("Batch_Summary.xlsx", summary_buffer.getvalue())
   
    zip_buffer.seek(0)
    return zip_buffer


# ============================================================
# 3. Streamlit App - FINAL VERSION
# ============================================================
st.set_page_config(page_title="Predictive Safety Scraper", layout="wide")


st.title("🚗 Predictive Safety HTML Scraper")
st.markdown("Upload HTML files containing crash frequency tables and extract Section 1 data.")


# Initialize session state
if 'step' not in st.session_state:
    st.session_state.step = 0  # 0=upload, 1=select, 3=results
if 'uploaded_files' not in st.session_state:
    st.session_state.uploaded_files = []
if 'file_names' not in st.session_state:
    st.session_state.file_names = []
if 'selected_files' not in st.session_state:
    st.session_state.selected_files = []
if 'results_data' not in st.session_state:
    st.session_state.results_data = {}
if 'inter_summary' not in st.session_state:
    st.session_state.inter_summary = pd.DataFrame()
if 'seg_summary' not in st.session_state:
    st.session_state.seg_summary = pd.DataFrame()
if 'processing_results' not in st.session_state:
    st.session_state.processing_results = []


# File uploader
uploaded_files = st.file_uploader(
    "Choose HTML files",
    type=['html', 'htm'],
    accept_multiple_files=True
)


# Handle new uploads
if uploaded_files and len(uploaded_files) != len(st.session_state.uploaded_files):
    st.session_state.uploaded_files = uploaded_files
    st.session_state.file_names = [f.name for f in uploaded_files]
    st.session_state.selected_files = []
    st.session_state.step = 1
    st.session_state.results_data = {}
    st.session_state.inter_summary = pd.DataFrame()
    st.session_state.seg_summary = pd.DataFrame()
    st.session_state.processing_results = []
    st.rerun()


# STEP 1: File Selection (Process starts immediately)
if st.session_state.step == 1 and st.session_state.file_names:
    st.subheader("📋 Select Files to Process")
    st.markdown("**Check files to process (all selected by default):**")
   
    # Create columns for buttons
    col_btn1, col_btn2, col_space = st.columns([1, 1, 3])
   
    with col_btn1:
        if st.button("✅ **Select All**", type="primary", use_container_width=True):
            st.session_state.selected_files = st.session_state.file_names.copy()
            st.rerun()
   
    with col_btn2:
        if st.button("❌ **Deselect All**", type="secondary", use_container_width=True):
            st.session_state.selected_files = []
            st.rerun()
   
    # Individual file checkboxes
    selected_files = []
   
    for i, filename in enumerate(st.session_state.file_names):
        # Get current selection state from session state for consistency
        current_key = f"file_{i}"
        if current_key not in st.session_state:
            # Initialize based on whether it's in selected_files
            st.session_state[current_key] = filename in st.session_state.selected_files
       
        is_selected = st.checkbox(
            f"📄 {filename}",
            key=current_key,
            value=st.session_state[current_key]
        )
        if is_selected:
            selected_files.append(filename)
   
    # Update session state with current selections
    st.session_state.selected_files = selected_files
   
    # Live count + selected files list
    col1, col2 = st.columns(2)
    with col1:
        st.info(f"**Selected: {len(selected_files)} / {len(st.session_state.file_names)}**")
    with col2:
        st.caption(f"*{len(st.session_state.file_names)} total files uploaded")
   
    if selected_files:
        with st.expander(f"📋 Selected Files ({len(selected_files)})", expanded=True):
            for filename in selected_files:
                st.markdown(f"✅ {filename}")
    else:
        st.warning("⚠️ No files selected")
   
    # DIRECT PROCESSING
    if st.button("🚀 Process Selected Files", type="primary", use_container_width=True):
        if selected_files:
            st.session_state.step = 3
            st.rerun()



# STEP 3: Processing + Results (COMBINED)
elif st.session_state.step == 3:
    # Process if not done yet
    if not st.session_state.processing_results:
        with st.spinner(f"🔄 Processing {len(st.session_state.selected_files)} files..."):
            inter_summary = pd.DataFrame()
            seg_summary = pd.DataFrame()
            results = []
           
            selected_uploaded_files = [
                f for f in st.session_state.uploaded_files
                if f.name in st.session_state.selected_files
            ]
           
            progress_bar = st.progress(0)
            status_text = st.empty()
           
            for i, uploaded_file in enumerate(selected_uploaded_files):
                filename = uploaded_file.name
                status_text.text(f"🔄 Processing: {filename} ({i+1}/{len(selected_uploaded_files)})")
               
                temp_path = f"temp_{i}_{filename.replace('/', '_')}"
                with open(temp_path, "wb") as f:
                    f.write(uploaded_file.getbuffer())
               
                try:
                    raw_df = extract_section1_df(temp_path)
                    consolidated = consolidate_tbl(raw_df.copy())
                    consolidated = rename_special_rows(consolidated)
                   
                    st.session_state.results_data[filename] = {
                        'raw': raw_df,
                        'consolidated': consolidated
                    }
                   
                    inter_summary, seg_summary = update_summary_tables(
                        filename, consolidated, inter_summary, seg_summary
                    )
                    results.append((filename, consolidated.shape[0], "✅ SUCCESS"))
                   
                except RuntimeError as e:
                    if "Could not find table with caption" in str(e):
                        results.append((filename, 0, "⚠️ NO Predicted Crash Frequencies and Rates TABLE"))
                    else:
                        results.append((filename, 0, f"❌ {str(e)[:40]}"))
                except Exception as e:
                    results.append((filename, 0, f"❌ ERROR: {str(e)[:40]}"))
               
                if os.path.exists(temp_path):
                    os.remove(temp_path)
               
                progress_bar.progress((i + 1) / len(selected_uploaded_files))
           
            st.session_state.inter_summary = inter_summary
            st.session_state.seg_summary = seg_summary
            st.session_state.processing_results = results
            status_text.success("✅ Processing complete!")
            st.rerun()
   
    # Show results
    st.success("🎉 Processing Complete!")
   
    st.subheader("📊 Processing Results")
    col1, col2, col3 = st.columns(3)
    with col1:
        successes = sum(1 for _, _, status in st.session_state.processing_results if "SUCCESS" in status)
        st.metric("✅ Successful", successes)
    with col2:
        st.metric("📁 Processed", len(st.session_state.processing_results))
    with col3:
        st.metric("⚠️ Issues", len(st.session_state.processing_results) - successes)
   
    results_df = pd.DataFrame(
        st.session_state.processing_results,
        columns=['File', 'Rows', 'Status']
    )
    st.dataframe(results_df, use_container_width=True)
   
    if st.button("🔄 Start Over", type="secondary"):
        for key in list(st.session_state.keys()):
            del st.session_state[key]
        st.rerun()
   
    st.divider()
   
    st.subheader("💾 Download Results")
    batch_zip = create_batch_zip(st.session_state.results_data, st.session_state.inter_summary, st.session_state.seg_summary)
    st.download_button(
        label="📦 Download Everything (ZIP)",
        data=batch_zip.getvalue(),
        file_name=f"Crash_Analysis_{len(st.session_state.results_data)}_files.zip",
        mime="application/zip",
        use_container_width=True
    )
   
    if not st.session_state.inter_summary.empty or not st.session_state.seg_summary.empty:
        st.subheader("📈 Summary Tables")
        col_inter, col_seg = st.columns(2)
       
        with col_inter:
            if not st.session_state.inter_summary.empty:
                st.write("**Intersection Crashes**")
                st.dataframe(append_total_row(st.session_state.inter_summary, 'Source File'), use_container_width=True)
       
        with col_seg:
            if not st.session_state.seg_summary.empty:
                st.write("**Segment Crashes**")
                st.dataframe(append_total_row(st.session_state.seg_summary, 'Source File'), use_container_width=True)


# Instructions
if st.session_state.step == 0:
    st.info("""
    **How to use:**
    1. 📤 Upload HTML files
    2. ✅ Check files to process (all selected by default)
    3. 🚀 Click "Process Selected Files"
    4. 💾 Download results
    """)